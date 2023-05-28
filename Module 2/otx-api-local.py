from __future__ import print_function
from OTXv2 import OTXv2
from tqdm import tqdm
import argparse
import hashlib
import IndicatorTypes
import pandas as pd
import openpyxl
import json
import requests
import textwrap
import pyodbc
import sys
import time
import re
import os

# Give the location of the file
path = "ioc.xlsx"
  
# Workbook object is created
wb_obj = openpyxl.load_workbook(path)  
sheet_obj = wb_obj.active
max_row = sheet_obj.max_row

# Your API key
API_KEY = '859db9fe4cdaf53c9e9ce1d3353caf89c47794e0b9121f469fad0ff299e9d00c'
OTX_SERVER = 'https://otx.alienvault.com/'
otx = OTXv2(API_KEY, server=OTX_SERVER)

parser = argparse.ArgumentParser(description='OTX CLI Example')
parser.add_argument('-ip', help='IP eg; 4.4.4.4', required=False)
parser.add_argument('-host',
                    help='Hostname eg; www.alienvault.com', required=False)
parser.add_argument(
    '-url', help='URL eg; http://www.alienvault.com', required=False)
parser.add_argument(
    '-hash', help='Hash of a file eg; 7b42b35832855ab4ff37ae9b8fa9e571', required=False)
parser.add_argument(
    '-file', help='Path to a file, eg; malware.exe', required=False)

args = vars(parser.parse_args())


def getValue(results, keys):
    if type(keys) is list and len(keys) > 0:

        if type(results) is dict:
            key = keys.pop(0)
            if key in results:
                return getValue(results[key], keys)
            else:
                return None
        else:
            if type(results) is list and len(results) > 0:
                return getValue(results[0], keys)
            else:
                return results
    else:
        return results

def hostname():

    for i in tqdm (range (100), desc="Analyzing Domains...", ascii=False, ncols=75):
        time.sleep(0.01)

    sheet_obj = wb_obj["domains"]
    max_row = sheet_obj.max_row

    for i in range(2, max_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 2)
        scan_value = cell_obj.value

        print(scan_value)
        
        response = otx.get_indicator_details_by_section(IndicatorTypes.DOMAIN, scan_value, 'general')

        # Return No if it's in the whitelist
        validation = getValue(response, ['validation'])
        if validation:
            threat = "No"
            print(threat)

        if not validation:
            pulses = getValue(response, ['pulse_info', 'pulses'])
            if pulses:
                threat = "Yes"
                print(threat)

            else:

                response = otx.get_indicator_details_by_section(IndicatorTypes.HOSTNAME, scan_value, 'general')
                # Return No if it's in the whitelist
                validation = getValue(response, ['validation'])
                if not validation:
                    pulses = getValue(response, ['pulse_info', 'pulses'])
                    if pulses:
                        threat = "Yes"
                        print(threat)
                    else:
                        threat = "No"
                        print(threat)

        sheet_obj.cell(row = i, column = 20).value = threat   

        wb_obj.save('ioc.xlsx') 



def url():

    for i in tqdm (range (100), desc="Analyzing URLs...", ascii=False, ncols=75):
        time.sleep(0.01)

    sheet_obj = wb_obj["urls"]
    max_row = sheet_obj.max_row

    for i in range(2, max_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 2)
        scan_value = cell_obj.value
        
        response = otx.get_indicator_details_by_section(IndicatorTypes.URL, scan_value, 'general')
        print(response)
        google = getValue( response, ['url_list', 'url_list', 'result', 'safebrowsing'])

        if google and 'response_code' in str(google):
            threat = "Yes"
            print(threat)

        else:
            threat = "No"
            print(threat)

        sheet_obj.cell(row = i, column = 15).value = threat   

        wb_obj.save('ioc.xlsx') 



def ips():

    for i in tqdm (range (100), desc="Analyzing IPs...", ascii=False, ncols=75):
        time.sleep(0.01)

    sheet_obj = wb_obj["ips"]
    max_row = sheet_obj.max_row

    for i in range(2, max_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 2)
        scan_value = cell_obj.value

        print(scan_value)
        
        response = otx.get_indicator_details_by_section(IndicatorTypes.IPv4, scan_value, 'general')

        # Return No if it's in the whitelist
        validation = getValue(response, ['validation'])
        if validation:
            threat = "No"
            print(threat)

        if not validation:
            pulses = getValue(response, ['pulse_info', 'pulses'])
            if pulses:
                threat = "Yes"
                print(threat)

            else:
                threat = "No"
                print(threat)

        sheet_obj.cell(row = i, column = 20).value = threat   

        wb_obj.save('ioc.xlsx') 



if args['host']:
    alerts = hostname(otx, args['host'])
    if len(alerts) > 0:
        print('Identified as potentially malicious')
        print(str(alerts))
    else:
        print('Unknown or not identified as malicious')

if args['ip']:
    alerts = ips(otx, args['ip'])
    if len(alerts) > 0:
        print('Identified as potentially malicious')
        print(str(alerts))

    else:
        print('Unknown or not identified as malicious')

if args['url']:
    alerts = url(otx, args['url'])
    if len(alerts) > 0:
        print('Identified as potentially malicious')
        print(str(alerts))
    else:
        print('Unknown or not identified as malicious')


def main():
    print("Starting OTX Analyzer..")

    for i in tqdm (range (100), desc="Loading...", ascii=False, ncols=75):
        time.sleep(0.01)

    hostname()
    # url()
    ips()




if __name__ == "__main__":
    main()

