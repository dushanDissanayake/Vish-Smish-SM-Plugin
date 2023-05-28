from __future__ import print_function
from tqdm import tqdm
import json
import hashlib
import requests
from virus_total_apis import PublicApi as VirusTotalPublicApi
import pandas as pd
import openpyxl
import os
import time
import sys

# Defining Public API KEY
API_KEY = 'd2400aee1b958962c56543c5efdf242d5b673fd399fb774ed35d3338502b5ce1'
vt = VirusTotalPublicApi(API_KEY)
  
# Give the location of the file
path = "ioc.xlsx"
  
# Workbook object is created
wb_obj = openpyxl.load_workbook(path)  
sheet_obj = wb_obj.active
max_row = sheet_obj.max_row


# Scanning IPs
def ipScan():

    for i in tqdm (range (100), desc="Analysing IPs...", ascii=False, ncols=75):
        time.sleep(0.01)

    sheet_obj = wb_obj["ips"]
    max_row = sheet_obj.max_row

    for i in range(2, max_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 2)
        scan_value = cell_obj.value

        print(scan_value)
        
        response = vt.get_ip_report(scan_value)

        time.sleep(15)

        print(response)

        detections = json.dumps(response['results']['positives'])
        totals = json.dumps(response['results']['total'])

        sheet_obj.cell(row = i, column = 3).value = detections
        sheet_obj.cell(row = i, column = 4).value = totals

        print(detections + " positive detections out of " + totals)    

        wb_obj.save('ioc.xlsx')
        

# Scanning Domains
def domainScan():
    print("\n")

    for i in tqdm (range (100), desc="Analysing Domains...", ascii=False, ncols=75):
        time.sleep(0.01)

    sheet_obj = wb_obj["domains"]
    max_row = sheet_obj.max_row

    for i in range(2, max_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 2)
        scan_value = cell_obj.value

        response = vt.get_domain_report(scan_value)
        time.sleep(15)

        reputation = json.dumps(response['results']['alphaMountain.ai category'])
        whois = json.dumps(response['results']['whois'])
        domain_siblings = json.dumps(response['results']['domain_siblings'])

        sheet_obj.cell(row = i, column = 3).value = reputation
        sheet_obj.cell(row = i, column = 5).value = whois
        sheet_obj.cell(row = i, column = 6).value = domain_siblings

        print(scan_value + " is a " + reputation + " website")

        wb_obj.save('ioc.xlsx')

# Scanning URLs
def urlScan():
    print("\n")

    for i in tqdm (range (100), desc="Analyzing URLs...", ascii=False, ncols=75):
        time.sleep(0.01)

    sheet_obj = wb_obj["urls"]
    max_row = sheet_obj.max_row

    for i in range(2, max_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 2)
        scan_value = cell_obj.value

        print(scan_value)
        
        response = vt.get_url_report(scan_value)
        # print(response)
        time.sleep(15)

        detections = json.dumps(response['results']['positives'])
        totals = json.dumps(response['results']['total'])

        print(detections + " positive detections out of " + totals)


        sheet_obj.cell(row = i, column = 3).value = detections
        sheet_obj.cell(row = i, column = 4).value = totals

        wb_obj.save('ioc.xlsx')


# def main():
#     print ('Starting VirusTotal Analyzing Module...')

#     for i in tqdm (range (100), desc="Loading...", ascii=False, ncols=75):
#         time.sleep(0.01)

#     while True:
#             try:
#                 choice = int(input("[?] What do you like to analyze? \
# 					\n1. IP\n2. File Hash\n3. Domain\n4. URL\n5. Exit\n\n"))
 
#             except ValueError:
#                 print ('[!] Please enter a valid number')
#                 continue      
    
#             if choice == 1:
#                 ipScan()
#                 break
#             if choice == 2:
#                 fileScan()
#                 break
#             if choice == 3:
#                 domainScan()
#                 break
#             if choice == 4:
#                 urlScan()
#                 break
#             if choice == 5:
#                 sys.exit(0)
#             else:
#                 print ('[!] Invalid Choice')


def main():
    print ('Starting VirusTotal Analyzing Module...')

    for i in tqdm (range (100), desc="Loading...", ascii=False, ncols=75):
        time.sleep(0.01)

    # ipScan()

    domainScan()

    urlScan()



if __name__ == "__main__":
    main()
