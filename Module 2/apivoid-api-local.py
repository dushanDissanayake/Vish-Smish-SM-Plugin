from tqdm import tqdm
import requests
import urllib.parse
import json
import sys
import os
import time
import openpyxl

apivoid_key = "b060c5c25454001e32afba4cf2ce8bea16ddd4b3";

# Give the location of the file
path = "ioc.xlsx"

# Workbook object is created
wb_obj = openpyxl.load_workbook(path)  


def apivoid_urlrep(key, url):
   try:
      r = requests.get(url='https://endpoint.apivoid.com/urlrep/v1/pay-as-you-go/?key='+key+'&url='+urllib.parse.quote(url))
      return json.loads(r.content.decode())
   except:
      return ""

def apivoid_parkeddomain(key, host):
   try:
      r = requests.get(url='https://endpoint.apivoid.com/parkeddomain/v1/pay-as-you-go/?key='+key+'&host='+host)
      return json.loads(r.content.decode())
   except:
      return ""

def apivoid_domainage(key, host):
   try:
      r = requests.get(url='https://endpoint.apivoid.com/domainage/v1/pay-as-you-go/?key='+key+'&host='+host)
      return json.loads(r.content.decode())
   except:
      return ""

def apivoid_iprep(key, ip):
   try:
      r = requests.get(url='https://endpoint.apivoid.com/iprep/v1/pay-as-you-go/?key='+key+'&ip='+ip)
      return json.loads(r.content.decode())
   except:
      return ""

def get_detection_engines(engines):
   list = "";
   for key, value in engines.items():
      if(bool(value['detected']) == 1):
         list+=str(value['engine'])+", "
   return list.rstrip(", ")

def apivoid_domainrep(key, host):
   try:
      r = requests.get(url='https://endpoint.apivoid.com/domainbl/v1/pay-as-you-go/?key='+key+'&host='+host)
      return json.loads(r.content.decode())
   except:
      return ""

def apivoid_sslrep(key, host):
   try:
      r = requests.get(url='https://endpoint.apivoid.com/sslinfo/v1/pay-as-you-go/?key='+key+'&host='+host)
      return json.loads(r.content.decode())
   except:
      return ""

def apivoid_emailrep(key, email):
   try:
      r = requests.get(url='https://endpoint.apivoid.com/emailverify/v1/pay-as-you-go/?key='+key+'&host='+email)
      return json.loads(r.content.decode())
   except:
      return ""


def scan_url():

    for i in tqdm (range (100), desc="Analyzing URLs...", ascii=False, ncols=75):
        time.sleep(0.01)

    print("\n")

    sheet_obj = wb_obj["urls"]
    max_row = sheet_obj.max_row

    for i in range(2, max_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 2)
        scan_value = cell_obj.value

        data = apivoid_urlrep(apivoid_key, scan_value)

        if(data):
            if(data.get('error')):
                print("Error: "+data['error'])
            else:
                # print("URL: "+str(scan_value))
                # print("Risk Score: "+str(data['data']['report']['risk_score']['result']))

                risk_score = json.dumps(str(data['data']['report']['risk_score']['result']))
                suspended = json.dumps(str(data['data']['report']['security_checks']['is_suspended_page']))
                blacklisted = json.dumps(str(data['data']['report']['security_checks']['is_domain_blacklisted']))
                phishing = json.dumps(str(data['data']['report']['security_checks']['is_phishing_heuristic']))
                sinkholed = json.dumps(str(data['data']['report']['security_checks']['is_sinkholed_domain']))
                ex_redirect = json.dumps(str(data['data']['report']['security_checks']['is_external_redirect']))
                pw_field = json.dumps(str(data['data']['report']['security_checks']['is_password_field']))
                crcard_field = json.dumps(str(data['data']['report']['security_checks']['is_credit_card_field']))
                url_short = json.dumps(str(data['data']['report']['site_category']['is_url_shortener']))
                free_host = json.dumps(str(data['data']['report']['site_category']['is_free_hosting']))


                sheet_obj.cell(row = i, column = 5).value = risk_score
                sheet_obj.cell(row = i, column = 6).value = suspended
                sheet_obj.cell(row = i, column = 7).value = blacklisted
                sheet_obj.cell(row = i, column = 8).value = phishing
                sheet_obj.cell(row = i, column = 9).value = sinkholed
                sheet_obj.cell(row = i, column = 10).value = ex_redirect
                sheet_obj.cell(row = i, column = 11).value = pw_field
                sheet_obj.cell(row = i, column = 12).value = crcard_field
                sheet_obj.cell(row = i, column = 13).value = url_short
                sheet_obj.cell(row = i, column = 14).value = free_host

                print("Risk Score = " + risk_score + ", Suspended Domain? = " + suspended + ", Blacklisted Domain? = " + blacklisted
                + ", Phishing Domain? = " + phishing + ", Sinhole Domain? = " + sinkholed + ", External Redirection? = " + ex_redirect + ", Contains Password Field = " + pw_field
                + ", Contains Credit Card Field = " + crcard_field + ", URL shortned? = " + url_short + ", Free Hosting? =" + free_host)   

                wb_obj.save('ioc.xlsx')

        else:
            print("Error: Request failed")



def scan_ip():

    for i in tqdm (range (100), desc="Analyzing IPs...", ascii=False, ncols=75):
        time.sleep(0.01)

    print("\n")

    sheet_obj = wb_obj["ips"]
    max_row = sheet_obj.max_row

    for i in range(2, max_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 2)
        scan_value = cell_obj.value

        print(scan_value)

        data = apivoid_iprep(apivoid_key, scan_value)

        if(data):
            if(data.get('error')):
                print("Error: "+data['error'])
            else:
                detection_count = json.dumps(str(data['data']['report']['blacklists']['detections']))
                is_proxy = json.dumps(str(data['data']['report']['anonymity']['is_proxy']))  
                is_tor = json.dumps(str(data['data']['report']['anonymity']['is_tor']))
                is_vpn = json.dumps(str(data['data']['report']['anonymity']['is_vpn']))
                is_hosting = json.dumps(str(data['data']['report']['anonymity']['is_hosting']))
               

                sheet_obj.cell(row = i, column = 3).value = detection_count
                sheet_obj.cell(row = i, column = 4).value = is_proxy
                sheet_obj.cell(row = i, column = 5).value = is_tor
                sheet_obj.cell(row = i, column = 6).value = is_vpn
                sheet_obj.cell(row = i, column = 7).value = is_hosting

                print("Detection Count = " + detection_count + ", Proxy? = " + is_proxy + ", Tor IP? = " + is_tor
                + ", VPN? = " + is_vpn + ", Hosting Provider? = " + is_hosting)   

                wb_obj.save('ioc.xlsx')

        else:
            print("Error: Request failed")



def scan_domain():

    for i in tqdm (range (100), desc="Analyzing Domains...", ascii=False, ncols=75):
        time.sleep(0.01)

    print("\n")

    sheet_obj = wb_obj["domains"]
    max_row = sheet_obj.max_row

    for i in range(2, max_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 2)
        scan_value = cell_obj.value

        data = apivoid_domainrep(apivoid_key, scan_value)
        parked_domain_data = apivoid_parkeddomain(apivoid_key, scan_value)
        domain_age_data = apivoid_domainage(apivoid_key, scan_value)

        if(data):

            if(data.get('error')):
                print("Error: "+data['error'])
            else:


                detection_count = json.dumps(str(data['data']['report']['blacklists']['detections']))
                is_free_host = json.dumps(str(data['data']['report']['category']['is_free_hosting']))
                is_url_shortner = json.dumps(str(data['data']['report']['category']['is_url_shortener']))
                is_free_dynamic_dns = json.dumps(str(data['data']['report']['category']['is_free_dynamic_dns']))
            

                sheet_obj.cell(row = i, column = 7).value = detection_count
                sheet_obj.cell(row = i, column = 8).value = is_free_host
                sheet_obj.cell(row = i, column = 9).value = is_url_shortner
                sheet_obj.cell(row = i, column = 10).value = is_free_dynamic_dns

                print("Detection Count = " + detection_count + ", Free Hosting? = " + is_free_host + ", URL Shortned? = " + is_url_shortner
                + ", Free Dynamic DNS? = " + is_free_dynamic_dns)   

                wb_obj.save('ioc.xlsx')

        else:
            print("Error: Request failed")

        if(parked_domain_data):

            if(parked_domain_data.get('error')):
                print("Error: "+data['error'])
            else:
                print(parked_domain_data)
                parked_status = json.dumps(str(parked_domain_data['data']['parked_domain']))
                print("Domain Parked? =" + parked_status)

                sheet_obj.cell(row = i, column = 12).value = parked_status

                wb_obj.save('ioc.xlsx')
        else:
            print("Error: Request failed")


        if(domain_age_data):
            if(domain_age_data.get('error')):
                print("Error: "+data['error'])
            else:
                print(domain_age_data)
                domain_age = json.dumps(str(domain_age_data['data']['domain_age_in_months']))
                print("Domain Age =" + domain_age + " Months")

                sheet_obj.cell(row = i, column = 13).value = domain_age


                wb_obj.save('ioc.xlsx')

        else:
            print("Error: Request failed")


def scan_email():


    for i in tqdm (range (100), desc="Analyzing Emails...", ascii=False, ncols=75):
        time.sleep(0.01)

    print("\n")

    sheet_obj = wb_obj["emails"]
    max_row = sheet_obj.max_row

    for i in range(2, max_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 2)
        scan_value = cell_obj.value

        data = apivoid_emailrep(apivoid_key, scan_value)

        if(data):

            if(data.get('error')):
                print("Error: "+data['error'])
            else:
                print(data)

                # detection_count = json.dumps(str(data['data']['report']['blacklists']['detections']))
            
                # sheet_obj.cell(row = i, column = 10).value = is_free_dynamic_dns

                print(sheet_obj.cell(row = i, column = 2).value)    

                wb_obj.save('ioc.xlsx')

        else:
            print("Error: Request failed")



def scan_ssl():

    sheet_obj = wb_obj.get_sheet_by_name("ssl")
    max_row = sheet_obj.max_row

    for i in range(2, max_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 2)
        scan_value = cell_obj.value

        data = apivoid_sslrep(apivoid_key, scan_value)

        if(data):
            if(data.get('error')):
                print("Error: "+data['error'])
            else:
                print('-')

                is_found = json.dumps(str(data['data']['certificate']['found']))
                fingerprint = json.dumps(str(data['data']['certificate']['fingerprint']))
                is_blacklisted = json.dumps(str(data['data']['certificate']['blacklisted']))
                valid_peer = json.dumps(str(data['data']['certificate']['valid_peer']))
                deprecated_issuer = json.dumps(str(data['data']['certificate']['deprecated_issuer']))
                name_match = json.dumps(str(data['data']['certificate']['name_match']))
                expired = json.dumps(str(data['data']['certificate']['expired']))
                valid = json.dumps(str(data['data']['certificate']['valid']))
                name = json.dumps(str(data['data']['certificate']['details']['subject']['name']))
                country = json.dumps(str(data['data']['certificate']['details']['subject']['country']))
                issuer_name = json.dumps(str(data['data']['certificate']['details']['issuer']['common_name']))
                valid_from = json.dumps(str(data['data']['certificate']['details']['validity']['valid_from']))
                valid_to = json.dumps(str(data['data']['certificate']['details']['validity']['valid_to']))
            

                sheet_obj.cell(row = i, column = 3).value = is_found
                sheet_obj.cell(row = i, column = 4).value = fingerprint
                sheet_obj.cell(row = i, column = 5).value = is_blacklisted
                sheet_obj.cell(row = i, column = 6).value = valid_peer
                sheet_obj.cell(row = i, column = 7).value = deprecated_issuer
                sheet_obj.cell(row = i, column = 8).value = name_match
                sheet_obj.cell(row = i, column = 9).value = expired
                sheet_obj.cell(row = i, column = 10).value = valid
                sheet_obj.cell(row = i, column = 11).value = name
                sheet_obj.cell(row = i, column = 12).value = country
                sheet_obj.cell(row = i, column = 13).value = issuer_name
                sheet_obj.cell(row = i, column = 14).value = valid_from
                sheet_obj.cell(row = i, column = 15).value = valid_to  


                print(sheet_obj.cell(row = i, column = 2).value)  

                wb_obj.save('ioc.xlsx')

        else:
            print("Error: Request failed")



# def main():
#     print ('-')
#     print ('-')
#     print ('Starting Artifact Reputation Analyzing Module...')

#     while True:
#             try:
#                 choice = int(input("[?] What do you like to analyze? \
# 					\n1. IP\n2. URL\n3. Domain\n4. SSL Certificate\n5. Email\n6. Exit\n\n"))

#             except ValueError:
#                 print ('[!] Please enter a valid number')
#                 continue      
    
#             if choice == 1:
#                 scan_ip()
#                 break
#             if choice == 2:
#                 scan_url()
#                 break
#             if choice == 3:
#                 scan_domain()
#                 break
#             if choice == 4:
#                 scan_ssl()
#                 break
#             if choice == 5:
#                 scan_email()
#                 break
#             if choice == 6:
#                 sys.exit(0)
#             else:
#                 print ('[!] Invalid Choice')


def main():

    print ('Starting ApiVoid Artifact Reputation Analysis Module...')

    # scan_ip()

    scan_url()

    scan_domain()

    # scan_email()



if __name__ == "__main__":
    main()