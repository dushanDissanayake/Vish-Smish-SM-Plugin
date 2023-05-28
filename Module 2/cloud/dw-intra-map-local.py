from __future__ import print_function
import json
import hashlib
import requests
from virus_total_apis import PublicApi as VirusTotalPublicApi
import pandas as pd
import openpyxl
import os
import time
import sys

# Defining Public API KEY
API_KEY = '0ba7ac4f8bebabf8c4773b58780a21a8c43ab284850b53c357d1a654fe847577'
vt = VirusTotalPublicApi(API_KEY)
  
# Give the location of the file
path = "../ioc-map.xlsx"
  
# Workbook object is created
wb_obj = openpyxl.load_workbook(path)  
sheet_obj = wb_obj.active
max_row = sheet_obj.max_row
  
# Apivoid configs
apivoid_key = "3af5da4d735edcd856838fd7fc51bf546193fa18";

def apivoid_sslrep(key, host):
   try:
      r = requests.get(url='https://endpoint.apivoid.com/sslinfo/v1/pay-as-you-go/?key='+key+'&host='+host)
      return json.loads(r.content.decode())
   except:
      return ""


# Scanning Downloads
def intramapFileScan():

    sheet_obj = wb_obj.get_sheet_by_name("downloads")
    max_row = sheet_obj.max_row
    no = "No"
    yes = "Yes"

    for i in range(2, 6):
        cell_obj = sheet_obj.cell(row = i, column = 11)
        scan_value = cell_obj.value
        print("Detection Count: ")
        print(scan_value)
        
        if (scan_value != 0):
            sheet_obj.cell(row = i, column = 15).value = no
            print("Flag Status: " + sheet_obj.cell(row = i, column = 15).value + "\n")
        else:
            cell_obj = sheet_obj.cell(row = i, column = 7)
            scan_value = cell_obj.value
            print ("Source Domain:  " + scan_value)
            response = vt.get_domain_report(scan_value)
            time.sleep(15)
            
            reputation = json.dumps(response['results']['alphaMountain.ai category'])

            print ("Domain Reputation:  " + reputation)
            
            if (reputation == "\"Malicious\""):
                sheet_obj.cell(row = i, column = 15).value = yes
                print ("Flag Status: " + sheet_obj.cell(row = i, column = 15).value + "\n")
            else:
                data = apivoid_sslrep(apivoid_key, scan_value)

                if(data):
                    if(data.get('error')):
                        print("Error: "+data['error'])
                    else:
                        is_found = json.dumps(str(data['data']['certificate']['found']))
                        print("Certificate Status:  " + is_found)

                        if (is_found == "\"False\""):
                            sheet_obj.cell(row = i, column = 15).value = yes
                            print ("Flag Status: " + sheet_obj.cell(row = i, column = 15).value + "\n")
                        else:
                            sheet_obj.cell(row = i, column = 15).value = no
                            print ("Flag Status: " + sheet_obj.cell(row = i, column = 15).value + "\n")
             
        
    wb_obj.save('ioc-map.xlsx')



def main():
    print('===========================================')
    print ('Starting Dear Watson Intracase Correlation Module...')

    while True:
            try:
                choice = int(input("[?] Initiate Intracase Correlation? \
					\n1. Yes\n2. Exit\n\n"))
            except ValueError:
                print ('[!] Please enter a valid number')
                continue      
    
            if choice == 1:
                intramapFileScan()
                break
            if choice == 2:
                sys.exit(0)
            else:
                print ('[!] Invalid Choice')


if __name__ == "__main__":
    main()
