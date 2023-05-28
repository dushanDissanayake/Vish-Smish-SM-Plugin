from __future__ import print_function
from tqdm import tqdm
import argparse
import hashlib
import pandas as pd
import openpyxl
import json
import requests
import textwrap
import pyodbc
import sys
import time
import re
import os

# Location of the file
path = "ioc.xlsx"
  
# Workbook object is created
wb_obj = openpyxl.load_workbook(path)  
sheet_obj = wb_obj.active
max_row = sheet_obj.max_row


def detect_ses_data():
    print ('Starting sensitive data identification module...')

    sheet_obj = wb_obj["bank"]

    cc_pattern = r'\b(4\d{3}|5[1-5]\d{2}|37\d{2})[-\s]?\d{4}[-\s]?\d{4}[-\s]?\d{4}\b'
    dob_pattern = r'\b(((19|20)\d\d)|([1-9]|0[1-9]|1[012])|(0[1-9]|[12][0-9]|3[01]))[- /.](([1-9]|0[1-9]|1[012])|(0[1-9]|[12][0-9]|3[01]))[- /.](((19|20)\d\d)|([1-9]|0[1-9]|1[012])|(0[1-9]|[12][0-9]|3[01]))\b'
    otp_pattern = r'\b(\d{8})|(\d{6})|(\d{4})|(G-\d{6})\b'
    ssn_pattern = r'(?<!\d)(?!000)(?!666)[0-8]\d{2}[- ]?(?!00)\d{2}[- ]?(?!0000)\d{4}(?!\d)'
    pw_pattern = r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{8,}$'
    iban_pattern = r'\b[a-zA-Z]{2}\d{2}[a-zA-Z0-9]{1,30}\b'
    address_pattern = r'(?:Road|St(?:reet)?|Avenue|Ave|Lane|Place|Blvd|Boulevard|Court|Ct|Drive|Dr|Parkway|Pkwy|Square|Sq|Way|Terrace|Ter|Circle|Cir|Highway|Hwy)$'

    for i in range(2, max_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 2)
        scan_value = cell_obj.value

        print(scan_value)
        print("\n")
    
        cc_match = re.search(cc_pattern, scan_value)
        dob_pattern = re.search(dob_pattern, scan_value)
        otp_pattern = re.search(otp_pattern, scan_value)
        ssn_pattern = re.search(ssn_pattern, scan_value)
        pw_pattern = re.search(pw_pattern, scan_value)
        iban_pattern = re.search(iban_pattern, scan_value)
        address_pattern = re.search(address_pattern, scan_value)

        print("Credit Card Number Matched:")
        print(cc_match)
        print("\n")

        print("Date of Birth Matched:")
        print(dob_pattern)
        print("\n")

        print("OTP Matched:")
        print(otp_pattern)
        print("\n")

        print("Social Security Number Matched:")
        print(ssn_pattern)
        print("\n")

        print("Password Matched:")
        print(pw_pattern)
        print("\n")

        print("Bank Account Number Matched:")
        print(iban_pattern)
        print("\n")

        print("Postal Address Matched:")
        print(address_pattern)
        print("\n")

        print("\n")

        return None


def main():
    detect_ses_data()


if __name__ == "__main__":
    main()


